import time
from progress_Bar import printProgressBar

from openpyxl import Workbook
wb = Workbook()

f = open("C:/Users/Kyle Perales/source/repos/exam-assembler/Parser/Resources/ocr_data.txt", "r")

# grab the active worksheet
ws = wb.active

tries = 1

# A List of Items
items = list(f)
l = len(items)
question = ''

# Initial call to print 0% progress
printProgressBar(0, l, prefix = 'Progress:', suffix = 'Complete', length = 50)
for i, item in enumerate(items):
    if item[0] == 'A' and item[1] == '.':
        Cindex = item.find('C.')
        #get A. option value
        Avalue = item[0:Cindex]
        key = 'B' + str(tries)
        ws[key] = Avalue

        #get C. option value
        endIndex = len(item)
        Cvalue = item[Cindex:endIndex]
        key = 'D' + str(tries)
        ws[key] = Cvalue
        
        #save question
        key = 'A' + str(tries)
        ws[key] = question
        question = ''

    elif item[0] == 'B' and item[1] == '.':
        
        Dindex = item.find('D.')
        #get B. option value
        Bvalue = item[0:Dindex]
        key = 'C' + str(tries)
        ws[key] = Bvalue

        #get D. option value
        endIndex = len(item)
        Dvalue = item[Dindex:endIndex]
        key = 'E' + str(tries)
        ws[key] = Dvalue
        
        tries += 1
    
    else: #get question
        question = question + item

    time.sleep(0.1)
    # Update Progress Bar
    printProgressBar(i + 1, l, prefix = 'Progress:', suffix = 'Complete', length = 50)

print("Done!")

# Save the file
wb.save("sorted_data.xlsx")