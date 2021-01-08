from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

for i in range(1,10):
    key = 'A' + str(i)
    ws[key] = i

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("test.xlsx")
